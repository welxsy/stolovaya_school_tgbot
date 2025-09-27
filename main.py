import pyodbc
import os
import re
import asyncio
import openpyxl
from openpyxl.styles import Font
from telegram import Update, ReplyKeyboardMarkup, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import ApplicationBuilder, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes
import nest_asyncio
from datetime import datetime, timedelta
import logging


# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger()

nest_asyncio.apply()

# Подключение к базе данных
conn = pyodbc.connect(
    "DRIVER={SQL Server};"
    "SERVER=localhost\\SQLEXPRESS;"
    "DATABASE=School12db;"
    "Trusted_Connection=yes;"
)
cursor = conn.cursor()

# Словарь для хранения временного списка выбранных учеников пользователями
user_selected_students = {}
BASE_FOLDER = "C:\\Users\\ASUS\\PycharmProjects\\еп\\user_files"


def delete_old_user_files(user_id: int, base_folder: str, days_old: int = 7):
    user_folder = os.path.join(base_folder, str(user_id))
    date_pattern = re.compile(r'\d{1,2}[А-Яа-я_]*_(\d{4})-(\d{2})-(\d{2})')
    today = datetime.today()
    deleted_files = []
    files_in_folder = os.listdir(user_folder)
    logger.info(f"Файлы в папке пользователя {user_id}: {files_in_folder}")

    for filename in files_in_folder:
        match = date_pattern.search(filename)
        if match:
            year, month, day = map(int, match.groups())
            file_date = datetime(year, month, day)
            if today - file_date > timedelta(days=days_old):
                file_path = os.path.join(user_folder, filename)
                os.remove(file_path)
                deleted_files.append(filename)
                logger.info(f"Файл {filename} удален.")
    if deleted_files:
        logger.info(f"Удалены старые файлы: {', '.join(deleted_files)}")
    else:
        logger.info(f"Нет старых файлов для удаления в папке пользователя {user_id}.")

    return deleted_files


async def view_previous_lists(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    folder_path = f"user_files/{user_id}"

    if not os.path.exists(folder_path) or not os.listdir(folder_path):
        await update.message.reply_text("❌ У вас нет сохраненных списков.")
        return

    files = os.listdir(folder_path)
    keyboard = [
        [InlineKeyboardButton(file, callback_data=f"view_file|{file}")] for file in files
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Выберите список, который хотите просмотреть:", reply_markup=reply_markup)


def create_user_folder(user_id):
    folder_path = f"user_files/{user_id}"
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    return folder_path


# Функция для создания или обновления Excel
def create_excel(class_name, students_list, user_id):
    folder_path = create_user_folder(user_id)
    today_date = datetime.now().strftime("%Y-%m-%d")

    for file in os.listdir(folder_path):
        if file.startswith(f"{class_name}_{today_date}"):
            os.remove(os.path.join(folder_path, file))

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["Фамилия ученика", "Имя ученика"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20

    for student in students_list:
        ws.append([student["LastName"], student["FirstName"]])

    ws.append(["Итого", len(students_list)])

    file_name = f"{class_name}_{today_date}.xlsx"
    file_path = os.path.join(folder_path, file_name)
    wb.save(file_path)
    return file_path


async def send_selected_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    file_name = query.data.split("|")[1]

    user_id = query.from_user.id
    folder_path = f"user_files/{user_id}"
    file_path = os.path.join(folder_path, file_name)

    if not os.path.exists(file_path):
        await query.message.reply_text("❌ Этот файл не существует.")
        return

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        students = []
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            last_name, first_name = row
            if last_name is None or first_name is None:
                continue
            if last_name == "Итого":
                continue
            students.append({"LastName": last_name, "FirstName": first_name})

        if not students:
            await query.message.reply_text("❌ В этом списке нет учеников.")
            return

        class_name, date_part = file_name.replace(".xlsx", "").split("_", 1)
        context.user_data["selected_class_name"] = class_name
        user_selected_students[user_id] = students

        student_text = "\n".join([f"- {s['LastName']} {s['FirstName']}" for s in students])
        message_text = f"📋 **Класс:** {class_name}\n📅 **Дата:** {date_part}\n\n**Ученики:**\n{student_text}"
        await query.message.reply_text(message_text, parse_mode="Markdown")
        keyboard = [["✏️ РЕДАКТИРОВАТЬ СПИСОК", "💾 СОХРАНИТЬ И ОТПРАВИТЬ СПИСОК"], ["🔙 Назад"]]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await query.message.reply_text("Вы можете отредактировать список или сохранить его как новый.",
                                       reply_markup=reply_markup)

    except Exception as e:
        logger.error(f"Ошибка при чтении файла: {e}")
        await query.message.reply_text("⚠️ Произошла ошибка при открытии списка.")


# Стартовое сообщение
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    user_name = update.message.from_user.full_name
    logger.info(f"Start command received from {user_name} (ID: {user_id})")

    # Проверка и удаление старых файлов перед началом работы
    deleted_files = delete_old_user_files(user_id, BASE_FOLDER, days_old=7)
    if deleted_files:
        logger.info(f"Удалены старые файлы для пользователя {user_name} (ID: {user_id}): {', '.join(deleted_files)}")

    keyboard = [["📋 МЕНЮ"], ["ℹ️ИНСТРУКЦИЯ"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text("👋 Добро пожаловать! Выберите действие:", reply_markup=reply_markup)


async def menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_name = update.message.from_user.full_name
    logger.info(f"Menu command received from {user_name} (ID: {update.message.from_user.id})")

    keyboard = [["📜 СОЗДАТЬ СПИСОК"], ["📂 ПРОШЛЫЕ СПИСКИ"], ["🔙 Назад"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text("📋 Главное меню", reply_markup=reply_markup)


async def info(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_name = update.message.from_user.full_name
    logger.info(f"Info command received from {user_name} (ID: {update.message.from_user.id})")

    await update.message.reply_text("ℹ️ Этот бот помогает создавать списки учеников!!!"
                                    "\n\n"
                                    "Главное меню:\n"
                                    "• Создать список — создайте новый список учеников.\n"
                                    "• Прошлые списки — посмотрите ранее сохранённые списки.\n"
                                    "• Назад — вернуться в предыдущее меню.\n"
                                    "\n"
                                    "Создание списка:\n"
                                    "• Выберите класс и добавьте учеников.\n"
                                    "• Вы можете просмотреть текущий список, редактировать его (добавить/удалить учеников) или сохранить в Excel.\n"
                                    "\n"
                                    "Редактирование:\n"
                                    "• Добавьте или удалите учеников из списка.\n"
                                    "\n"
                                    "Просмотр старых списков:\n"
                                    "• Перейдите в раздел Прошлые списки для доступа к ранее сохранённым.\n"
                                    "\n"
                                    "После выбора класса вы сможете добавить учеников в список, отредактировать его и затем нажать кнопку 'Сохранить и отправить.'"
                                    "\n\n"
                                    "!!!!! ДЛЯ ТОГО ЧТОБЫ ОТРЕДАКТИРОВАТЬ УЖЕ ОТПРАВЛЕННЫЙ СПИСОК, ПРОСТО СОЗДАЙТЕ НОВЫЙ. ОТПРАВЛЕН БУДЕТ ПОСЛЕДНИЙ СОЗДАННЫЙ СПИСОК.")


# Выбор класса
async def choose_class(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cursor.execute("SELECT ClassID, ClassName FROM Classes")
    classes = cursor.fetchall()

    keyboard = [
        [InlineKeyboardButton(c.ClassName, callback_data=f"class_{c.ClassID}")] for c in classes
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text("Выберите класс:", reply_markup=reply_markup)


async def send_students_buttons(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    class_id = query.data.split("_")[1]

    cursor.execute("SELECT ClassName FROM Classes WHERE ClassID=?", class_id)
    class_name = cursor.fetchone()[0]
    context.user_data["selected_class_name"] = class_name

    cursor.execute("SELECT LastName, FirstName FROM Students WHERE ClassID=?", class_id)
    students = cursor.fetchall()
    if not students:
        await query.message.reply_text("❌ В этом классе нет учеников.")  # Отладочное сообщение
        return

    user_id = query.from_user.id
    user_selected_students[user_id] = []

    keyboard = [
        [InlineKeyboardButton(f"{s[0]} {s[1]}", callback_data=f"student_{s[0]}_{s[1]}")] for s in students
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.message.reply_text(f"Вы выбрали класс: {class_name}. Теперь выберите учеников:",
                                   reply_markup=reply_markup)

    action_keyboard = ReplyKeyboardMarkup(
        [["📄 ПРОСМОТРЕТЬ ТЕКУЩИЙ СПИСОК"], ["💾 СОХРАНИТЬ И ОТПРАВИТЬ СПИСОК"], ["✏️ РЕДАКТИРОВАТЬ СПИСОК"], ["🔙 Назад"]],
        resize_keyboard=True)
    await query.message.reply_text("Выберите действие:", reply_markup=action_keyboard)


# Добавление выбранных учеников
async def student_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    selected_student = query.data.split("_")[1:]
    if user_id not in user_selected_students:
        user_selected_students[user_id] = []
    student = {"LastName": selected_student[0], "FirstName": selected_student[1]}
    if student not in user_selected_students[user_id]:
        user_selected_students[user_id].append(student)

    user_name = query.from_user.full_name
    logger.info(f"Student {selected_student[0]} {selected_student[1]} selected by {user_name} (ID: {user_id})")

    await query.message.reply_text(f"Вы добавили ученика: {selected_student[0]} {selected_student[1]}")


# Просмотр текущего списка
async def view_current_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    if user_id not in user_selected_students or not user_selected_students[user_id]:
        await update.message.reply_text("❌ Список пуст.")
        return
    student_list = "\n".join([f"{s['LastName']} {s['FirstName']}" for s in user_selected_students[user_id]])
    await update.message.reply_text(f"📄 Текущий список:\n{student_list}")


async def edit_student_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [InlineKeyboardButton("➕ Добавить ученика", callback_data="edit_add_student")],
        [InlineKeyboardButton("🗑️ Удалить ученика", callback_data="edit_remove_student")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Что вы хотите сделать со списком?", reply_markup=reply_markup)


async def handle_edit_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    action = query.data

    if action == "edit_add_student":
        class_name = context.user_data.get("selected_class_name")
        if not class_name:
            await query.message.reply_text("❌ Класс не выбран.")
            return

        cursor.execute("SELECT ClassID FROM Classes WHERE ClassName=?", class_name)
        result = cursor.fetchone()
        if result:
            class_id = result[0]


            cursor.execute("SELECT LastName, FirstName FROM Students WHERE ClassID=?", class_id)
            students = cursor.fetchall()

            if not students:
                await query.message.reply_text("❌ В этом классе нет учеников.")
                return

            keyboard = [
                [InlineKeyboardButton(f"{s[0]} {s[1]}", callback_data=f"student_{s[0]}_{s[1]}")] for s in students
            ]

            reply_markup = InlineKeyboardMarkup(keyboard)
            await query.message.reply_text(f"Выберите ученика для добавления в класс {class_name}:",
                                           reply_markup=reply_markup)
    elif action == "edit_remove_student":
        user_id = query.from_user.id
        if user_id not in user_selected_students or not user_selected_students[user_id]:
            await query.message.reply_text("❌ Список пуст.")
            return

        student_list = user_selected_students[user_id]
        keyboard = [
            [InlineKeyboardButton(f"Удалить {s['LastName']} {s['FirstName']}",
                                  callback_data=f"remove_student_{s['LastName']}_{s['FirstName']}")]
            for s in student_list
        ]

        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.reply_text("Выберите ученика для удаления:", reply_markup=reply_markup)


async def remove_student(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    student_list = user_selected_students.get(user_id, [])
    if not student_list:
        await query.message.reply_text("❌ Список пуст. Некого удалять.")
        return

    student_to_remove = query.data.split("_")[2:]
    student_to_remove = {"LastName": student_to_remove[0], "FirstName": student_to_remove[1]}

    if student_to_remove in student_list:
        student_list.remove(student_to_remove)
        user_selected_students[user_id] = student_list
        total_students = len(student_list)
        student_list.append({"LastName": "Итого", "FirstName": total_students})

        await query.message.reply_text(
            f"Ученик {student_to_remove['LastName']} {student_to_remove['FirstName']} удален.")
    else:
        await query.message.reply_text("❌ Этот ученик не найден в списке.")


async def generate_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    if user_id not in user_selected_students or not user_selected_students[user_id]:
        await update.message.reply_text("❌ Не выбраны ученики для создания списка.")
        return

    class_name = context.user_data.get("selected_class_name")
    file_path = create_excel(class_name, user_selected_students[user_id], user_id)
    with open(file_path, "rb") as file:
        await update.message.reply_document(document=file)

    admin_chat_id = 6129878481
    with open(file_path, "rb") as file:
        await update.message._bot.send_document(chat_id=admin_chat_id, document=file)

    user_selected_students[user_id] = []


# Логирование сообщений
async def log_messages(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_name = update.message.from_user.full_name
    logger.info(f"Message received from {user_name} (ID: {update.message.from_user.id}): {update.message.text}")

    text = update.message.text
    if text == "📋 МЕНЮ":
        await menu(update, context)
    elif text == "📜 СОЗДАТЬ СПИСОК":
        await choose_class(update, context)
    elif text == "ℹ️ИНСТРУКЦИЯ":
        await info(update, context)
    elif text == "📄 ПРОСМОТРЕТЬ ТЕКУЩИЙ СПИСОК":
        await view_current_list(update, context)
    elif text == "💾 СОХРАНИТЬ И ОТПРАВИТЬ СПИСОК":
        await generate_excel(update, context)
    elif text == "✏️ РЕДАКТИРОВАТЬ СПИСОК":
        await edit_student_list(update, context)
    elif text == "📂 ПРОШЛЫЕ СПИСКИ":
        await view_previous_lists(update, context)
    elif text == "🔙 Назад":
        await start(update, context)



async def main():
    app = ApplicationBuilder().token("7472359286:AAHVC0ibd9FenDWSMoesj7YCydR4VbC_NIA").build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("menu", menu))
    app.add_handler(CommandHandler("info", info))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, log_messages))
    app.add_handler(CallbackQueryHandler(send_students_buttons, pattern="^class_"))
    app.add_handler(CallbackQueryHandler(student_selected, pattern="^student_"))
    app.add_handler(CallbackQueryHandler(remove_student, pattern="^remove_student_"))
    app.add_handler(CallbackQueryHandler(handle_edit_menu, pattern="^edit_"))
    app.add_handler(CallbackQueryHandler(send_selected_file, pattern="^view_file\\|"))

    await app.run_polling()


if __name__ == '__main__':
    asyncio.run(main())